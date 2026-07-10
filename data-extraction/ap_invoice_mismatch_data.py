"""
AP Invoice Mismatch Data Extractor  v1
Finds records where RecordType or VendorName was changed by the customer:
  - Sheet 1: RecordType mismatches  (AAI vs Customer)
  - Sheet 2: VendorName mismatches  (AAI vs Customer)

Includes ALL columns from ap_invoice_data.py + extra mismatch-specific columns
(customer entity_name, AAI/Customer RecordType, Customer vendorName from edits).
"""

import requests
import pandas as pd
import sys, time, getpass, json, re, os
from openpyxl import load_workbook

# ===== CONFIGURATION =====
METABASE_REGULAR_URL = "https://metabase.auditoria.ai"
METABASE_ENT_URL     = "https://metabase-ent1.auditoria.ai"
USERNAME_REGULAR     = "venu.gopal@auditoria.ai"
USERNAME_ENT         = "venu.gopal@auditoria.ai"
OUTPUT_FILENAME      = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     "AP_Invoice_Mismatch_Report.xlsx")

# ── Date range ──
QUERY_DATE_FROM = '2026-06-01 11:00:00'
QUERY_DATE_TO   = '2026-06-29 11:00:00'

# All tenant IDs to fetch
ALL_TENANT_IDS = [
    '683032998341775360',
    '665247456933969920',
    '796485641444532224',
    '546005944371253248'
]

MAX_CSV_CELL_CHARS   = 500
ROWS_PER_FILE        = 5000
QUERY_TIMEOUT        = 600
PRECHECK_TIMEOUT     = 15
SILENT_QUERY_TIMEOUT = 30
BULK_FETCH_TIMEOUT   = 300

# ===== SQL QUERY TEMPLATE =====
# Fetches all records where EITHER recordType OR vendorName was changed.
# Driven from sor.wk_result_edits (swre) so we capture the edit diffs.
SQL_QUERY_TEMPLATE = """
SELECT
    x.`Document ID`,
    x.created_at,
    x.updated_at,
    x.`Edit Message ID`,
    x.`Message ID`,
    x.`Tenant Name`,
    x.`Tenant ID`,
    x.`customer entity_name`,
    x.`On UI`,
    x.`Reason for Dismissal`,
    x.`Written`,
    x.`Manual`,
    x.`Final Record Type`,
    x.`Original Record Type`,
    x.`AAI RecordType`,
    x.`Customer RecordType`,
    x.`Invoice #`,
    x.vendorid,
    x.vendorname,
    x.OriginalVendorName,
    x.`Customer vendorName`,
    x.ExtractedVendorName,
    x.NormalizedVendorName,
    x.VendorNameReason,
    x.doc_classification_json,
    x.vendorname_json,

    CASE WHEN x.NormalizedVendorName IS NOT NULL
              AND x.OriginalVendorName LIKE CONCAT('%%', x.NormalizedVendorName, '%%')
         THEN 'Yes' ELSE 'No' END AS normalized_matched_with_OriginalvendorName,

    CASE WHEN x.NormalizedVendorName IS NOT NULL
              AND x.vendorname LIKE CONCAT('%%', x.NormalizedVendorName, '%%')
         THEN 'Yes' ELSE 'No' END AS normalized_matched_with_finalvendorName,

    CASE WHEN x.ExtractedVendorName IS NOT NULL
              AND x.OriginalVendorName LIKE CONCAT('%%', x.ExtractedVendorName, '%%')
         THEN 'Yes' ELSE 'No' END AS Extracted_matched_with_OriginalvendorName,

    CASE WHEN x.ExtractedVendorName IS NOT NULL
              AND x.vendorname LIKE CONCAT('%%', x.ExtractedVendorName, '%%')
         THEN 'Yes' ELSE 'No' END AS Extraction_matched_with_finalvendorName,

    x.`Vendor Match Status`,
    x.attachmentFileName,
    x.extractedFileS3Location,
    x.originalAttachmentFileName,
    x.S3Location,
    x.recipient_email,
    x.sender_email
FROM (
    SELECT
        a.id AS "Document ID",
        swre.created_at,
        swre.updated_at,
        swre.message_id AS "Edit Message ID",
        JSON_UNQUOTE(JSON_EXTRACT(a.session, '$.messages[0]')) AS "Message ID",
        t.name AS 'Tenant Name',
        t.id AS "Tenant ID",
        se.entity_name AS 'customer entity_name',
        CASE a.active
            WHEN 0 THEN 'Dismissed'
            WHEN 1 THEN 'Active'
            ELSE 'Unknown'
        END AS 'On UI',
        JSON_UNQUOTE(JSON_EXTRACT(a.record, '$.dismissedComment')) AS 'Reason for Dismissal',
        CASE a.write_status
            WHEN 0 THEN 'No'
            WHEN 1 THEN 'Written'
            ELSE 'Unknown'
        END AS "Written",
        CASE a.isManualEntry
            WHEN 0 THEN 'Bot'
            WHEN 1 THEN 'Manual'
            ELSE 'Unknown'
        END AS 'Manual',
        JSON_UNQUOTE(JSON_EXTRACT(a.record, '$.recordType')) AS 'Final Record Type',
        JSON_UNQUOTE(JSON_EXTRACT(a.session, '$.originalRecordType')) AS 'Original Record Type',
        swre.original_json->>'$.recordType' AS 'AAI RecordType',
        swre.final_json->>'$.recordType' AS 'Customer RecordType',
        JSON_UNQUOTE(JSON_EXTRACT(a.record, '$.invoiceNumber')) AS 'Invoice #',
        JSON_UNQUOTE(JSON_EXTRACT(a.record, '$.vendorId')) AS vendorid,
        JSON_UNQUOTE(JSON_EXTRACT(a.record, '$.vendorName')) AS vendorname,
        swre.original_json->>'$.vendorName' AS OriginalVendorName,
        swre.final_json->>'$.vendorName' AS 'Customer vendorName',

        COALESCE(
            (SELECT jt.extracted_value FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',extracted_value VARCHAR(500) PATH '$.Value')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1),
            (SELECT jt.extracted_value FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*][*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',extracted_value VARCHAR(500) PATH '$.Value')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1)
        ) AS ExtractedVendorName,

        COALESCE(
            (SELECT jt.normalized_value FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',normalized_value VARCHAR(500) PATH '$.NormalizedValue')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1),
            (SELECT jt.normalized_value FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*][*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',normalized_value VARCHAR(500) PATH '$.NormalizedValue')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1)
        ) AS NormalizedVendorName,

        COALESCE(
            (SELECT jt.reason FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',reason VARCHAR(255) PATH '$.Reason')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1),
            (SELECT jt.reason FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*][*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',reason VARCHAR(255) PATH '$.Reason')) jt
             WHERE LOWER(jt.field_name)='vendorname' LIMIT 1)
        ) AS VendorNameReason,

        COALESCE(
            CAST(JSON_EXTRACT(mia.intent,'$.s3Attachments[0].DocumentData[0].Data.doc_classification') AS CHAR(10000)),
            '[]'
        ) AS doc_classification_json,

        COALESCE(
            (SELECT JSON_ARRAYAGG(jt.full_object) FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',full_object JSON PATH '$')) jt
             WHERE LOWER(jt.field_name)='vendorname'),
            (SELECT JSON_ARRAYAGG(jt.full_object) FROM JSON_TABLE(mia.intent,'$.extractedInfo.Fields[*][*]'
                COLUMNS(field_name VARCHAR(100) PATH '$.Name',full_object JSON PATH '$')) jt
             WHERE LOWER(jt.field_name)='vendorname'),
            JSON_ARRAY()
        ) AS vendorname_json,

        CASE
            WHEN JSON_UNQUOTE(JSON_EXTRACT(a.record,'$.vendorName'))=swre.original_json->>'$.vendorName' THEN 'Vendor Match'
            WHEN swre.original_json->>'$.vendorName' IS NULL THEN 'No Original Data'
            ELSE 'Vendor Mismatch'
        END AS 'Vendor Match Status',

        JSON_UNQUOTE(JSON_EXTRACT(a.session,'$.attachments[0].filename'))            AS attachmentFileName,
        JSON_UNQUOTE(JSON_EXTRACT(a.session,'$.attachments[0].key'))                 AS extractedFileS3Location,
        JSON_UNQUOTE(JSON_EXTRACT(a.session,'$.originalAttachment.filename'))        AS originalAttachmentFileName,
        JSON_UNQUOTE(JSON_EXTRACT(a.session,'$.originalAttachment.attchment[0].key')) AS S3Location,
        mia.intent->>'$.To'   AS recipient_email,
        mia.intent->>'$.From' AS sender_email

    FROM sor.wk_result_edits swre
    LEFT JOIN sor.wk_inst_result a ON swre.wk_result_id = a.id
    LEFT JOIN tenant.tenant t ON t.id = swre.tenant_id AND t.type='customer' AND t.sandbox='0'
    LEFT JOIN conversation.message m ON m.id = JSON_UNQUOTE(JSON_EXTRACT(a.session,'$.messages[0]'))
    LEFT JOIN conversation.message_intent_audit mia ON mia.message_id = m.id AND mia.app_def_code='VIDE'
    LEFT JOIN sor.entity se ON se.entity_id = swre.final_json->>'$.aaiEntityId'
    WHERE
        a.write_status = 1
        AND swre.created_at >= 'DATE_FROM_PLACEHOLDER'
        AND swre.created_at <= 'DATE_TO_PLACEHOLDER'
        AND swre.tenant_id = 'TENANT_PLACEHOLDER'
        AND (
            swre.original_json->>'$.recordType' != swre.final_json->>'$.recordType'
            OR swre.original_json->>'$.vendorName' != a.record ->> '$.vendorName'
            OR swre.original_json->>'$.aaiEntityId' != swre.final_json->>'$.aaiEntityId'
        )
) x
ORDER BY x.updated_at DESC
"""


def _elapsed(t): return f"{time.time()-t:.1f}s"

def sanitize_for_sql(v):
    if pd.isna(v): return ''
    return str(v).replace("'", "''")

def _sql_for_tenant(tenant_id):
    return (SQL_QUERY_TEMPLATE
            .replace('DATE_FROM_PLACEHOLDER', QUERY_DATE_FROM)
            .replace('DATE_TO_PLACEHOLDER',   QUERY_DATE_TO)
            .replace('TENANT_PLACEHOLDER',    tenant_id))

def _truncate_for_csv(df, mx=MAX_CSV_CELL_CHARS):
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].apply(
                lambda x: (s := str(x) if pd.notna(x) else '') and
                          ((s[:mx] + '...') if len(s) > mx else s) or '')
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Auth & DB listing
# ─────────────────────────────────────────────────────────────────────────────

def authenticate(url, username, password):
    try:
        print(f"  Authenticating {url} ...", end=" ", flush=True)
        r = requests.post(f"{url}/api/session",
                          json={"username": username, "password": password}, timeout=30)
        if r.status_code == 200:
            print("OK"); return r.json()['id']
        print(f"FAIL (HTTP {r.status_code})"); return None
    except Exception as e:
        print(f"FAIL ({e})"); return None


def list_databases(url, token):
    try:
        r = requests.get(f"{url}/api/database",
                         headers={"X-Metabase-Session": token}, timeout=30)
        if r.status_code == 200:
            dbs = r.json().get('data', [])
            print(f"\n  Databases on {url}:")
            for db in dbs:
                print(f"   ID {db['id']:<5}  {db['name']:<35}  engine={db.get('engine','?')}")
            return dbs
        return []
    except Exception as e:
        print(f"FAIL ({e})"); return []


# ─────────────────────────────────────────────────────────────────────────────
# Core query executor
# ─────────────────────────────────────────────────────────────────────────────

# Snowflake ID columns: keep as text end-to-end. Building the DataFrame with dtype=object (not the
# default per-column inference) prevents a NULL turning an ID column into float64 — which would
# corrupt the 18-digit IDs at read time, before any export.
_ID_COLS = ('Document ID', 'Tenant ID', 'Message ID', 'Edit Message ID', 'vendorid')

def _build_df(rows, cols):
    df = pd.DataFrame(rows, columns=cols, dtype=object)
    for c in _ID_COLS:
        if c in df.columns:
            df[c] = df[c].apply(lambda v: '' if v is None or (isinstance(v, float) and pd.isna(v)) else str(v))
    return df

def _parse_df(result):
    if 'cols' in result and 'rows' in result:
        return _build_df(result['rows'], [c['name'] for c in result['cols']])
    d = result.get('data', {})
    if 'cols' in d and 'rows' in d:
        return _build_df(d['rows'], [c['name'] for c in d['cols']])
    return None


def execute_query(url, token, db_id, sql, timeout=QUERY_TIMEOUT, verbose=True):
    hdr     = {"X-Metabase-Session": token, "Content-Type": "application/json"}
    payload = {"database": db_id, "type": "native", "native": {"query": sql.strip()}}
    t0      = time.time()
    if verbose:
        print(f"   Querying DB {db_id} ...", flush=True)
    try:
        r = requests.post(f"{url}/api/dataset", headers=hdr, json=payload, timeout=timeout)
    except requests.exceptions.Timeout:
        if verbose: print(f"   Timed out after {timeout}s")
        return None
    except Exception as e:
        if verbose: print(f"   Error: {e}")
        return None

    if r.status_code == 200:
        result = r.json(); df = _parse_df(result)
        if df is None:
            if verbose: print("   Could not parse response"); return None
        if verbose:
            print(f"   {len(df):,} rows  [{_elapsed(t0)}]")
        return df

    if r.status_code == 202:
        body = r.json(); df = _parse_df(body)
        if df is not None:
            if verbose: print(f"   {len(df):,} rows  [{_elapsed(t0)}]")
            return df
        for via in body.get('via', []):
            if via.get('status') == 'failed':
                if verbose: print("   Query failed"); return None
        qk = body.get('query_key') or body.get('query_uuid') or body.get('uuid')
        if qk:
            if verbose: print("   Polling ...", flush=True)
            for i in range(1, 61):
                time.sleep(5)
                if verbose: print(f"      poll {i}/60  [{_elapsed(t0)}]", flush=True)
                try:
                    pr = requests.get(f"{url}/api/dataset/query_key/{qk}",
                                      headers=hdr, timeout=30)
                    if pr.status_code == 200:
                        df = _parse_df(pr.json())
                        if df is not None:
                            if verbose: print(f"   {len(df):,} rows  [{_elapsed(t0)}]")
                            return df
                except Exception: continue
            if verbose: print("   Polling timed out"); return None
        if verbose: print(f"   202 with no data - DB {db_id} likely wrong schema")
        return None

    if verbose:
        try:   msg = r.json().get('message', r.text[:150])
        except: msg = r.text[:150]
        print(f"   HTTP {r.status_code}: {msg[:150]}")
    return None


def execute_query_silent(url, token, db_id, sql, timeout=SILENT_QUERY_TIMEOUT):
    return execute_query(url, token, db_id, sql, timeout=timeout, verbose=False)


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Find which DB has each tenant  (fast COUNT probe)
# ─────────────────────────────────────────────────────────────────────────────

def find_tenant_db(tenant_id, instances):
    probe = f"""
    SELECT COUNT(*) AS cnt FROM sor.wk_result_edits swre
    LEFT JOIN sor.wk_inst_result a ON swre.wk_result_id = a.id
    WHERE a.write_status = 1
      AND swre.created_at >= '{QUERY_DATE_FROM}'
      AND swre.created_at <= '{QUERY_DATE_TO}'
      AND swre.tenant_id = '{tenant_id}'
      AND (
          swre.original_json->>'$.recordType' != swre.final_json->>'$.recordType'
          OR swre.original_json->>'$.vendorName' != a.record ->> '$.vendorName'
      )
    """
    for inst in instances:
        db_ids = inst['db_ids']
        for db_id in db_ids:
            r = execute_query_silent(inst['url'], inst['token'], db_id,
                                     probe, timeout=PRECHECK_TIMEOUT)
            if r is not None and not r.empty:
                cnt = r.iloc[0, 0]
                try: cnt = int(cnt)
                except: cnt = 0
                if cnt > 0:
                    return {
                        'url':   inst['url'],
                        'token': inst['token'],
                        'db_id': db_id,
                        'name':  inst['name'],
                        'rows':  cnt,
                    }
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Fetch data for a single tenant from its known DB
# ─────────────────────────────────────────────────────────────────────────────

def fetch_tenant_data(tenant_id, db_info):
    sql = _sql_for_tenant(tenant_id)
    print(f"   Running mismatch query for tenant {tenant_id} on {db_info['name']} DB {db_info['db_id']} ...")
    df = execute_query(db_info['url'], db_info['token'], db_info['db_id'],
                       sql, timeout=QUERY_TIMEOUT, verbose=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Main orchestration: per-tenant find + fetch
# ─────────────────────────────────────────────────────────────────────────────

def fetch_all_tenants(instances):
    results = []
    grand_t0 = time.time()

    print(f"\n{'='*70}")
    print(f"  STEP 1 - Find which DB has mismatch data for each tenant")
    print(f"{'='*70}")

    tenant_db_map = {}
    for tid in ALL_TENANT_IDS:
        print(f"\n  Probing for tenant {tid} ...", end=" ", flush=True)
        t_probe = time.time()
        info = find_tenant_db(tid, instances)
        if info:
            tenant_db_map[tid] = info
            print(f"-> found on {info['name']} DB {info['db_id']}  "
                  f"(~{info['rows']} mismatch rows)  [{_elapsed(t_probe)}]")
        else:
            print(f"-> NO MISMATCHES found on any instance  [{_elapsed(t_probe)}]")

    print(f"\n{'-'*70}")
    print(f"  Tenant -> DB map:")
    for tid, info in tenant_db_map.items():
        print(f"    {tid}  ->  {info['name']} DB {info['db_id']}  (~{info['rows']} mismatches)")
    missing = [t for t in ALL_TENANT_IDS if t not in tenant_db_map]
    if missing:
        for t in missing:
            print(f"    {t}  ->  no mismatch data found")

    print(f"\n{'='*70}")
    print(f"  STEP 2 - Fetch mismatch data per tenant")
    print(f"{'='*70}")

    for i, (tid, db_info) in enumerate(tenant_db_map.items(), 1):
        print(f"\n  [{i}/{len(tenant_db_map)}] Tenant {tid}")
        t0 = time.time()
        df = fetch_tenant_data(tid, db_info)
        if df is not None and not df.empty:
            df['Data_Source'] = db_info['name']
            results.append(df)
            print(f"   {len(df):,} rows fetched  [{_elapsed(t0)}]")
        else:
            print(f"   No data returned for tenant {tid}  [{_elapsed(t0)}]")

    if not results:
        print("\nNo mismatch data from any tenant"); return None, tenant_db_map

    combined = pd.concat(results, ignore_index=True)

    if 'Document ID' in combined.columns:
        before = len(combined)
        combined = combined.drop_duplicates(subset=['Document ID'], keep='first')
        if before - len(combined):
            print(f"\nRemoved {before-len(combined):,} duplicates")

    if 'updated_at' in combined.columns:
        combined = combined.sort_values('updated_at', ascending=False)

    print(f"\n{'='*70}")
    print(f"Total: {len(combined):,} mismatch rows  [{_elapsed(grand_t0)}]")
    print(f"\nSummary:")
    if 'Tenant ID' in combined.columns:
        for tid, cnt in combined['Tenant ID'].value_counts().items():
            src = combined[combined['Tenant ID'] == tid]['Data_Source'].iloc[0]
            print(f"   Tenant {tid}: {cnt:,} rows  ({src})")
    missing_final = [t for t in ALL_TENANT_IDS
                     if t not in set(combined.get('Tenant ID', pd.Series()).unique())]
    if missing_final:
        print(f"\n   Still missing (no mismatches in date range): {missing_final}")

    return combined, tenant_db_map


# ─────────────────────────────────────────────────────────────────────────────
# Post-processing  (same as ap_invoice_data.py)
# ─────────────────────────────────────────────────────────────────────────────

def update_vendor_names_from_json(df):
    print("\n" + "-"*70)
    print("Updating vendor names from vendorname_json ...")
    updated = 0
    for idx, row in df.iterrows():
        orig = str(row.get('OriginalVendorName', '') or '').strip()
        raw  = str(row.get('vendorname_json', '[]'))
        if not orig or orig == 'nan' or raw == '[]': continue
        try:
            entries = json.loads(raw)
            if not isinstance(entries, list): continue
            for e in entries:
                if not isinstance(e, dict): continue
                val  = str(e.get('Value', '') or '').strip()
                norm = str(e.get('NormalizedValue', '') or '').strip()
                ol   = orig.lower()
                if (val and val.lower() in ol) or (norm and norm.lower() in ol):
                    if val  and val  != 'nan': df.at[idx, 'ExtractedVendorName']  = val
                    if norm and norm != 'nan': df.at[idx, 'NormalizedVendorName'] = norm
                    updated += 1; break
        except Exception: pass
    print(f"   Updated {updated}/{len(df)} rows")
    return df


def format_datetime_columns(df):
    for col in ['created_at', 'updated_at']:
        if col not in df.columns: continue
        def fmt(val):
            if pd.isna(val): return ''
            try:
                dt = pd.to_datetime(val) if isinstance(val, str) else val
                return dt.strftime('%B %d, %Y, %-I:%M %p')
            except: return str(val)
        df[col] = df[col].apply(fmt)
    return df


def format_dates_in_json_columns(df):
    json_cols = [
        'sor_hints_value_normalized_matched_with_OriginalvendorName',
        'sor_hints_value_Extracted_matched_with_OriginalvendorName',
        'sor_master_value_normalized_matched_with_OriginalvendorName',
        'sor_master_value_Extracted_matched_with_OriginalvendorName',
        'Systemhints_value_for_Normalized_VendorName',
        'Systemhints_value_for_Extracted_VendorName',
    ]
    def fmt_d(s):
        if not s or s == 'None': return ''
        try: return pd.to_datetime(s).strftime('%B %d, %Y, %-I:%M %p')
        except: return s
    for col in json_cols:
        if col not in df.columns: continue
        def fmt_j(raw):
            if not raw or raw == '[]': return raw
            try:
                data = json.loads(raw)
                if isinstance(data, list):
                    for e in data:
                        if isinstance(e, dict):
                            for k in ('created_at', 'updated_at'):
                                if k in e: e[k] = fmt_d(e[k])
                return json.dumps(data)
            except: return raw
        df[col] = df[col].apply(fmt_j)
    return df


def _like_pattern_to_regex(v):
    parts = []
    for ch in v:
        if   ch == '%': parts.append('.*')
        elif ch == '_': parts.append('.')
        else:           parts.append(re.escape(ch))
    return '.*' + ''.join(parts) + '.*'


def _like_match_series(series, v):
    if series is None or len(series) == 0:
        return pd.Series([False] * (0 if series is None else len(series)))
    return series.astype(str).str.contains(_like_pattern_to_regex(v),
                                            regex=True, case=False, na=False)


def _bulk_fetch_for_tenant(tid, db_info, tables_needed):
    out = {'hints': None, 'master_sor': None, 'system_hints': None}
    safe_tid = sanitize_for_sql(tid)
    if 'hints' in tables_needed:
        sql = (f"SELECT active_hints, data->>'$.name' AS vendor_name, "
               f"data->>'$.acceptableNames' AS acceptable_name, "
               f"data->>'$.address.address1' AS address, "
               f"tenant_id, created_at, updated_at, data AS _match_data "
               f"FROM sor.hints WHERE tenant_id='{safe_tid}' ORDER BY id")
        out['hints'] = execute_query_silent(db_info['url'], db_info['token'],
                                             db_info['db_id'], sql,
                                             timeout=BULK_FETCH_TIMEOUT)
    if 'master_sor' in tables_needed:
        sql = (f"SELECT created_at, updated_at, coalesced_name, status, "
               f"data->>'$.address1' AS address1, data->>'$.address2' AS address2, "
               f"tenant_id "
               f"FROM sor.master_sor WHERE tenant_id='{safe_tid}' "
               f"AND status='active' ORDER BY id")
        out['master_sor'] = execute_query_silent(db_info['url'], db_info['token'],
                                                  db_info['db_id'], sql,
                                                  timeout=BULK_FETCH_TIMEOUT)
    if 'system_hints' in tables_needed:
        sql = (f"SELECT created_at, created_by, updated_at, updated_by, "
               f"extracted_name, initial_vendor_name, final_vendor_name, "
               f"write_status, tenant_id "
               f"FROM sor.system_hints WHERE tenant_id='{safe_tid}' ORDER BY id")
        out['system_hints'] = execute_query_silent(db_info['url'], db_info['token'],
                                                    db_info['db_id'], sql,
                                                    timeout=BULK_FETCH_TIMEOUT)
    return out


def _match_hints(hints_df, v):
    if hints_df is None or hints_df.empty: return None
    mask = _like_match_series(hints_df['_match_data'], v)
    matched = hints_df[mask].head(10).copy()
    if matched.empty: return None
    matched = matched.drop(columns=['_match_data'])
    matched.insert(3, 'search_value_used', v)
    return matched


def _match_master_sor(ms_df, v):
    if ms_df is None or ms_df.empty: return None
    mask = _like_match_series(ms_df['coalesced_name'], v)
    matched = ms_df[mask].head(10).copy()
    if matched.empty: return None
    matched.insert(4, 'search_value_used', v)
    return matched


def _match_system_hints(sh_df, v):
    if sh_df is None or sh_df.empty: return None
    mask = (_like_match_series(sh_df['extracted_name'], v)
            | _like_match_series(sh_df['initial_vendor_name'], v)
            | _like_match_series(sh_df['final_vendor_name'], v))
    matched = sh_df[mask].head(10).copy()
    if matched.empty: return None
    matched.insert(8, 'search_value_used', v)
    return matched


def run_post_processing_queries(df, instances, tenant_db_map):
    print("\n" + "="*70)
    print("Post-Processing  (bulk per-tenant fetch + local match)")
    print("="*70)

    POST_COLS = [
        'sor_hints_value_normalized_matched_with_OriginalvendorName',
        'sor_hints_value_Extracted_matched_with_OriginalvendorName',
        'sor_master_value_normalized_matched_with_OriginalvendorName',
        'sor_master_value_Extracted_matched_with_OriginalvendorName',
        'Systemhints_value_for_Normalized_VendorName',
        'Systemhints_value_for_Extracted_VendorName',
    ]
    for c in POST_COLS: df[c] = '[]'

    if 'Tenant ID' not in df.columns:
        print("No 'Tenant ID' column; nothing to do."); return df

    tenants_in_df = [str(t) for t in df['Tenant ID'].dropna().unique()
                     if str(t) not in ('', 'nan')]

    def _tables_needed(tid):
        reasons = df.loc[df['Tenant ID'].astype(str) == tid, 'VendorNameReason'].astype(str)
        need = set()
        if reasons.str.contains('withHints', na=False).any():
            need.update({'hints', 'master_sor'})
        if reasons.str.contains('withLLM', na=False).any():
            need.add('master_sor')
        if reasons.str.contains('fromSystemHints', na=False).any():
            need.update({'system_hints', 'master_sor'})
        return need

    def _find_db_fallback(tid):
        probe = f"SELECT tenant_id FROM sor.hints WHERE tenant_id='{sanitize_for_sql(tid)}' LIMIT 1"
        for inst in instances:
            for db_id in inst['db_ids']:
                r = execute_query_silent(inst['url'], inst['token'], db_id, probe)
                if r is not None and not r.empty:
                    return {'url': inst['url'], 'token': inst['token'],
                            'db_id': db_id, 'name': inst['name']}
        return None

    print("\nBulk-fetching sor.* per tenant ...")
    tenant_data = {}
    for tid in tenants_in_df:
        need = _tables_needed(tid)
        if not need:
            tenant_data[tid] = {'hints': None, 'master_sor': None, 'system_hints': None}
            continue
        db_info = tenant_db_map.get(tid) or _find_db_fallback(tid)
        if not db_info:
            print(f"   Tenant {tid}: no DB found - post-proc columns stay '[]'")
            tenant_data[tid] = {'hints': None, 'master_sor': None, 'system_hints': None}
            continue
        print(f"   Tenant {tid} -> {db_info['name']} DB {db_info['db_id']}  "
              f"(fetch: {sorted(need)})")
        tenant_data[tid] = _bulk_fetch_for_tenant(tid, db_info, need)
        for tbl, tdf in tenant_data[tid].items():
            if tdf is not None:
                print(f"      {tbl}: {len(tdf):,} rows")

    total = len(df); stats = {'ok': 0, 'empty': 0, 'err': 0}; t0 = time.time()
    print(f"\nMatching {total:,} rows locally ...")

    for rn, (idx, row) in enumerate(df.iterrows(), 1):
        if rn == 1 or rn == total or rn % 500 == 0:
            el = time.time() - t0; rate = rn/el if el else 0
            print(f"  [{rn:>{len(str(total))}}/{total}] {rn/total*100:5.1f}%  "
                  f"rate={rate:.0f}/s  ok={stats['ok']} empty={stats['empty']} "
                  f"err={stats['err']}", flush=True)

        tid    = str(row.get('Tenant ID', '') or '').strip()
        reason = str(row.get('VendorNameReason', '') or '')
        extr_v = str(row.get('ExtractedVendorName', '') or '').strip()
        norm_v = str(row.get('NormalizedVendorName', '') or '').strip()

        if not reason or reason == 'nan': continue
        data = tenant_data.get(tid)
        if not data: continue

        def assign(col, matched_df):
            try:
                if matched_df is None or matched_df.empty:
                    stats['empty'] += 1
                else:
                    df.at[idx, col] = matched_df.to_json(orient='records')
                    stats['ok'] += 1
            except Exception as e:
                stats['err'] += 1
                print(f"      {col}: {str(e)[:100]}")

        if 'withHints' in reason:
            for v, ch, cm in [
                (norm_v, 'sor_hints_value_normalized_matched_with_OriginalvendorName',
                         'sor_master_value_normalized_matched_with_OriginalvendorName'),
                (extr_v, 'sor_hints_value_Extracted_matched_with_OriginalvendorName',
                         'sor_master_value_Extracted_matched_with_OriginalvendorName'),
            ]:
                if not v or v == 'nan': continue
                assign(ch, _match_hints(data['hints'], v))
                assign(cm, _match_master_sor(data['master_sor'], v))

        elif 'withLLM' in reason:
            for v, cm in [
                (norm_v, 'sor_master_value_normalized_matched_with_OriginalvendorName'),
                (extr_v, 'sor_master_value_Extracted_matched_with_OriginalvendorName'),
            ]:
                if not v or v == 'nan': continue
                assign(cm, _match_master_sor(data['master_sor'], v))

        elif 'fromSystemHints' in reason:
            for v, cs, cm in [
                (extr_v, 'Systemhints_value_for_Extracted_VendorName',
                         'sor_master_value_Extracted_matched_with_OriginalvendorName'),
                (norm_v, 'Systemhints_value_for_Normalized_VendorName',
                         'sor_master_value_normalized_matched_with_OriginalvendorName'),
            ]:
                if not v or v == 'nan': continue
                assign(cs, _match_system_hints(data['system_hints'], v))
                assign(cm, _match_master_sor(data['master_sor'], v))

    print(f"\n{'='*70}")
    print(f"POST-PROCESSING: ok={stats['ok']} empty={stats['empty']} err={stats['err']}")
    for col in POST_COLS:
        n = (df[col] != '[]').sum()
        print(f"  {n:>5}/{len(df)} ({n/len(df)*100:4.1f}%)  {col}")
    print("="*70)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Split into RecordType / VendorName mismatch sheets
# ─────────────────────────────────────────────────────────────────────────────

def split_mismatches(df):
    rt_mask = df.apply(
        lambda r: (str(r.get('AAI RecordType', '') or '').strip() !=
                   str(r.get('Customer RecordType', '') or '').strip())
                  and str(r.get('AAI RecordType', '') or '').strip() not in ('', 'None', 'nan')
                  and str(r.get('Customer RecordType', '') or '').strip() not in ('', 'None', 'nan'),
        axis=1
    )
    vn_mask = df.apply(
        lambda r: (str(r.get('OriginalVendorName', '') or '').strip() !=
                   str(r.get('vendorname', '') or '').strip())
                  and str(r.get('OriginalVendorName', '') or '').strip() not in ('', 'None', 'nan')
                  and str(r.get('vendorname', '') or '').strip() not in ('', 'None', 'nan'),
        axis=1
    )

    df_rt = df[rt_mask].copy().reset_index(drop=True)
    df_vn = df[vn_mask].copy().reset_index(drop=True)

    print(f"\nMismatch breakdown:")
    print(f"   RecordType mismatches : {len(df_rt):,} rows")
    print(f"   VendorName mismatches : {len(df_vn):,} rows")
    both = rt_mask & vn_mask
    print(f"   Both mismatched       : {both.sum():,} rows")

    return df_rt, df_vn


# ─────────────────────────────────────────────────────────────────────────────
# Excel output  (two sheets: RecordType_Mismatch + VendorName_Mismatch)
# ─────────────────────────────────────────────────────────────────────────────

def save_to_excel(df_rt, df_vn, filename, truncate=True):
    if (df_rt is None or df_rt.empty) and (df_vn is None or df_vn.empty):
        print("Nothing to save"); return False
    if not filename.endswith('.xlsx'):
        filename = filename.replace('.csv', '.xlsx')

    ID_COLS = ['Document ID', 'Tenant ID', 'Message ID', 'Edit Message ID', 'vendorid']
    def id_str(x):
        # Snowflake IDs are 18 digits; float() loses precision above 2^53 and corrupts ~all of
        # them, so NEVER route an ID through float(). Normalise to the exact digit string instead.
        if x is None: return ''
        if isinstance(x, float):
            if pd.isna(x): return ''
            return '%.0f' % x  # last-resort integral formatting (dtype=str at read keeps IDs off this)
        s = str(x).strip()
        if s in ('', 'nan', 'NaN', 'None', '<NA>'): return ''
        return s[:-2] if s.endswith('.0') and s[:-2].isdigit() else s

    sheets = []
    if df_rt is not None and not df_rt.empty:
        to_save = _truncate_for_csv(df_rt) if truncate else df_rt.copy()
        for col in ID_COLS:
            if col in to_save.columns: to_save[col] = to_save[col].apply(id_str)
        sheets.append(('RecordType_Mismatch', to_save))

    if df_vn is not None and not df_vn.empty:
        to_save = _truncate_for_csv(df_vn) if truncate else df_vn.copy()
        for col in ID_COLS:
            if col in to_save.columns: to_save[col] = to_save[col].apply(id_str)
        sheets.append(('VendorName_Mismatch', to_save))

    if not sheets:
        print("No mismatch data to save"); return False

    total_rows = sum(len(s[1]) for s in sheets)
    print(f"\nSaving {total_rows:,} rows -> {filename}  ({len(sheets)} sheet(s)) ...")

    with pd.ExcelWriter(filename, engine='openpyxl') as w:
        for sname, sdf in sheets:
            sdf.to_excel(w, sheet_name=sname, index=False)
            print(f"   Sheet '{sname}': {len(sdf):,} rows x {len(sdf.columns)} cols")

    wb = load_workbook(filename)
    for ws in wb.worksheets:
        hdr = {cell.value: cell.column for cell in ws[1]}
        for cn in ID_COLS:
            if cn not in hdr: continue
            cl = ws.cell(1, hdr[cn]).column_letter
            for r in range(2, ws.max_row + 1):
                c = ws[f'{cl}{r}']
                if c.value: c.number_format = '@'
    wb.save(filename)
    print(f"\nSaved: {filename}")
    for sname, sdf in sheets:
        print(f"   {sname}: {len(sdf):,} rows x {len(sdf.columns)} cols")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st = time.time()
    print("\n" + "="*70)
    print("  AP INVOICE MISMATCH DATA EXTRACTOR  (RecordType + VendorName)")
    print("="*70)
    print(f"\nTenants : {ALL_TENANT_IDS}")
    print(f"Date range: {QUERY_DATE_FROM}  ->  {QUERY_DATE_TO}")
    print(f"Filter: write_status=1, recordType OR vendorName mismatch\n")

    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             '..', 'predictive_coding', '.env')
    env_vars = {}
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1); env_vars[k.strip()] = v.strip()
        print("Loaded credentials from .env")

    instances = []

    print(f"\nRegular Metabase  ({METABASE_REGULAR_URL})")
    pw_r = env_vars.get('METABASE_REG_PASSWORD') or getpass.getpass("   Password: ")
    t_r  = authenticate(METABASE_REGULAR_URL, USERNAME_REGULAR, pw_r)
    if not t_r: print("Cannot continue"); sys.exit(1)
    dbs_r  = list_databases(METABASE_REGULAR_URL, t_r)
    ids_r  = [d['id'] for d in dbs_r if any(k in d['name'].lower() for k in ('sor', 'shard'))]
    if not ids_r: ids_r = [d['id'] for d in dbs_r]
    print(f"DB IDs: {ids_r}")
    instances.append({'name': 'Regular', 'url': METABASE_REGULAR_URL,
                      'token': t_r, 'db_ids': ids_r})

    print(f"\nENT Metabase  ({METABASE_ENT_URL})")
    pw_e = env_vars.get('METABASE_ENT_PASSWORD') or getpass.getpass("   Password: ")
    t_e  = authenticate(METABASE_ENT_URL, USERNAME_ENT, pw_e)
    if t_e:
        dbs_e  = list_databases(METABASE_ENT_URL, t_e)
        ids_e  = [d['id'] for d in dbs_e if any(k in d['name'].lower() for k in ('sor', 'shard'))]
        if not ids_e: ids_e = [d['id'] for d in dbs_e]
        if 6  in ids_e: ids_e.remove(6);  ids_e.insert(0, 6)
        if 18 not in ids_e: ids_e.append(18)
        print(f"DB IDs: {ids_e}")
        instances.append({'name': 'ENT', 'url': METABASE_ENT_URL,
                          'token': t_e, 'db_ids': ids_e})
    else:
        print("ENT auth failed")

    combined, tenant_db_map = fetch_all_tenants(instances)
    if combined is None or combined.empty:
        print("\nNo mismatch data found"); sys.exit(1)

    combined = update_vendor_names_from_json(combined)
    combined = format_datetime_columns(combined)
    combined = run_post_processing_queries(combined, instances, tenant_db_map)
    combined = format_dates_in_json_columns(combined)

    df_rt, df_vn = split_mismatches(combined)

    save_to_excel(df_rt, df_vn, OUTPUT_FILENAME, truncate=False)

    print(f"\n{'='*70}")
    print(f"Done  - {len(combined):,} total mismatch rows  |  {_elapsed(st)}")
    print(f"   RecordType mismatches : {len(df_rt):,}")
    print(f"   VendorName mismatches : {len(df_vn):,}")
    print(f"   File: {OUTPUT_FILENAME}")
    print("="*70 + "\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted"); sys.exit(0)
    except Exception as e:
        import traceback; traceback.print_exc(); sys.exit(1)
