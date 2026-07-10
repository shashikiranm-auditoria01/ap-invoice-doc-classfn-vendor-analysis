import asyncio
import csv
import io
import json
import logging
import os
import time
from typing import AsyncGenerator, Optional

import httpx
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Metabase Realtime Query UI")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

METABASE_INSTANCES = {
    "regular": "https://metabase.auditoria.ai",
    "enterprise": "https://metabase-ent1.auditoria.ai",
}

# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------

class LoginRequest(BaseModel):
    instance: str  # "regular" or "enterprise"
    username: str
    password: str


class QueryRequest(BaseModel):
    instance: str
    session_token: str
    database_id: int
    sql: str


class ExportRequest(BaseModel):
    columns: list[str]
    rows: list[list]
    format: str = "excel"  # "excel" or "csv"


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def get_base_url(instance: str) -> str:
    url = METABASE_INSTANCES.get(instance)
    if not url:
        raise HTTPException(status_code=400, detail=f"Unknown instance: {instance}. Use 'regular' or 'enterprise'.")
    return url


def metabase_headers(session_token: str) -> dict:
    return {
        "Content-Type": "application/json",
        "X-Metabase-Session": session_token,
    }


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.post("/api/login")
async def login(req: LoginRequest):
    base_url = get_base_url(req.instance)
    async with httpx.AsyncClient(verify=False, timeout=30) as client:
        try:
            resp = await client.post(
                f"{base_url}/api/session",
                json={"username": req.username, "password": req.password},
                headers={"Content-Type": "application/json"},
            )
        except httpx.ConnectError as e:
            raise HTTPException(status_code=503, detail=f"Cannot reach {base_url}: {e}")
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Connection timed out")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if resp.status_code != 200:
        raise HTTPException(status_code=resp.status_code, detail=f"Metabase error: {resp.text}")

    data = resp.json()
    token = data.get("id") or data.get("token")
    if not token:
        raise HTTPException(status_code=500, detail="No session token in Metabase response")

    return {"session_token": token, "instance": req.instance, "base_url": base_url}


# ---------------------------------------------------------------------------
# Databases
# ---------------------------------------------------------------------------

@app.get("/api/databases")
async def list_databases(instance: str, session_token: str):
    base_url = get_base_url(instance)
    async with httpx.AsyncClient(verify=False, timeout=30) as client:
        try:
            resp = await client.get(
                f"{base_url}/api/database",
                headers=metabase_headers(session_token),
                params={"include_tables": "true"},
            )
        except httpx.ConnectError as e:
            raise HTTPException(status_code=503, detail=str(e))
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Timed out fetching databases")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="Session expired or invalid")
    if resp.status_code != 200:
        raise HTTPException(status_code=resp.status_code, detail=resp.text)

    data = resp.json()
    # Metabase returns either a list or {"data": [...]}
    databases = data if isinstance(data, list) else data.get("data", data)

    result = []
    for db in databases:
        tables = db.get("tables", [])
        result.append({
            "id": db["id"],
            "name": db["name"],
            "engine": db.get("engine", ""),
            "tables": [
                {"id": t["id"], "name": t["name"], "schema": t.get("schema", "")}
                for t in tables
            ],
        })
    return result


# ---------------------------------------------------------------------------
# Query — SSE streaming
# ---------------------------------------------------------------------------

async def stream_query(instance: str, session_token: str, database_id: int, sql: str) -> AsyncGenerator[str, None]:
    base_url = get_base_url(instance)
    payload = {
        "database": database_id,
        "type": "native",
        "native": {"query": sql},
    }

    yield _sse("status", {"message": "Connecting to Metabase…"})

    try:
        async with httpx.AsyncClient(verify=False, timeout=300) as client:
            async with client.stream(
                "POST",
                f"{base_url}/api/dataset",
                json=payload,
                headers=metabase_headers(session_token),
            ) as resp:
                if resp.status_code == 401:
                    yield _sse("error", {"message": "Session expired. Please log in again."})
                    return
                if resp.status_code == 403:
                    yield _sse("error", {"message": "Permission denied for this database."})
                    return
                if resp.status_code not in (200, 202):
                    body = await resp.aread()
                    try:
                        err = json.loads(body).get("message", body.decode())
                    except Exception:
                        err = body.decode()
                    yield _sse("error", {"message": f"Metabase error ({resp.status_code}): {err}"})
                    return

                yield _sse("status", {"message": "Query running…"})

                # Collect full response body (Metabase returns one big JSON blob)
                chunks = []
                async for chunk in resp.aiter_bytes(chunk_size=65536):
                    chunks.append(chunk)

                raw = b"".join(chunks)

        # Parse the response
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as e:
            yield _sse("error", {"message": f"Failed to parse Metabase response: {e}"})
            return

        # Handle Metabase error responses embedded in 200 OK
        if data.get("error"):
            yield _sse("error", {"message": str(data["error"])})
            return

        result_data = data.get("data", data)
        cols = result_data.get("cols") or result_data.get("results_metadata", {}).get("columns", [])
        native_rows = result_data.get("rows") or []

        columns = [c.get("display_name") or c.get("name", f"col_{i}") for i, c in enumerate(cols)]

        yield _sse("columns", {"columns": columns})
        yield _sse("status", {"message": f"Streaming {len(native_rows):,} rows…"})

        # Stream rows in batches for progressive rendering
        BATCH = 500
        total = len(native_rows)
        for i in range(0, total, BATCH):
            batch = native_rows[i : i + BATCH]
            yield _sse("rows", {"rows": batch, "offset": i, "total": total})
            await asyncio.sleep(0)  # yield control so SSE flushes

        yield _sse("done", {"total": total, "columns": columns})

    except httpx.ConnectError as e:
        yield _sse("error", {"message": f"Cannot reach {base_url}: {e}"})
    except httpx.TimeoutException:
        yield _sse("error", {"message": "Query timed out (300s limit exceeded)"})
    except Exception as e:
        logger.exception("Unexpected error during query streaming")
        yield _sse("error", {"message": f"Unexpected error: {e}"})


def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


@app.post("/api/query/stream")
async def query_stream(req: QueryRequest):
    return StreamingResponse(
        stream_query(req.instance, req.session_token, req.database_id, req.sql),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.post("/api/export/excel")
async def export_excel(req: ExportRequest):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    # Header row styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, col_name in enumerate(req.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, row in enumerate(req.rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns (cap at 60)
    for col_idx in range(1, len(req.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(req.columns[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"metabase_results_{int(time.time())}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/export/csv")
async def export_csv(req: ExportRequest):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(req.columns)
    writer.writerows(req.rows)
    buf.seek(0)

    filename = f"metabase_results_{int(time.time())}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Serve frontend
# ---------------------------------------------------------------------------

static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

    @app.get("/")
    async def root():
        return FileResponse(os.path.join(static_dir, "index.html"))
else:
    @app.get("/")
    async def root():
        return JSONResponse({"status": "Metabase API backend running", "docs": "/docs"})
